"""Excel parser using openpyxl (offline)."""
from pathlib import Path

from ..models import DocumentFormat, ParsedDocument
from .base import BaseParser


class XlsxParser(BaseParser):
    supported_formats = [DocumentFormat.XLSX]

    def parse(self, path: Path) -> ParsedDocument:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"=== Sheet: {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        lines.append("\t".join(cells))
            raw_text = "\n".join(lines)
        except ImportError:
            raw_text = f"[openpyxl not installed — cannot parse {path.name}]"
        except Exception as exc:  # noqa: BLE001
            raw_text = f"[XLSX parse error: {exc}]"

        return ParsedDocument(
            source_path=str(path),
            format=DocumentFormat.XLSX,
            title=path.stem,
            raw_text=raw_text,
        )
